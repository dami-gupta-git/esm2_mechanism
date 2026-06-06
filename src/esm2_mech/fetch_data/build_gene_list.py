"""
Build gene_list.tsv — step 1 of the gene universe pipeline.

Merges Gerasimavicius et al. (DiseaseMech_Stability_VEPS.xlsx) and G2P
(AllG2P.csv) into a single gene list with canonical mechanism labels.

Inputs : data/downloads/DiseaseMech_Stability_VEPS.xlsx
         data/downloads/AllG2P.csv
Output : data/gene_list.tsv
         Columns: gene, mechanism, uniprot_id, source, g2p_disagrees

Usage:
    python -m esm2_mech.fetch_data.build_gene_list
"""

from __future__ import annotations

import csv
import functools
from collections import Counter
from pathlib import Path

import openpyxl
import pandas as pd

from esm2_mech.utils.paths import ALL_G2P_FILE, DISEASE_MECH_STABILITY_VEPS_FILE, GENE_LIST_TSV

print = functools.partial(print, flush=True)

XLSX = DISEASE_MECH_STABILITY_VEPS_FILE
G2P_CSV = ALL_G2P_FILE

G2P_MECH_MAP = {
    "loss of function": "LOF",
    "gain of function": "GOF",
    "dominant negative": "DN",
}
G2P_CONFIDENCE_KEEP = {"definitive", "strong"}

LEGACY_GENE_ALIASES: list[dict] = [
    {
        "gene": "C12ORF65",
        "mechanism": "AR",
        "uniprot_id": "Q9H3J6",
        "source": "gerasimavicius",
        "g2p_disagrees": "",
    },
    {
        "gene": "C19ORF12",
        "mechanism": "LOF",
        "uniprot_id": "",
        "source": "g2p",
        "g2p_disagrees": "",
    },
]

CV_MECH_MAP = {
    "AR, Het": "AR",
    "AR": "AR",
    "HI": "HI",
    "GOF": "GOF",
    "DN": "DN",
    "Unknown": "Unknown",
}


def _load_functional_protein_class(wb) -> dict[str, str]:
    ws = wb["Functional_protein_class"]
    out: dict[str, str] = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        gene, _inh, mech = row[0], row[1], row[2]
        if gene:
            out[gene] = mech or "Unknown"
    print(f"Functional_protein_class: {len(out)} genes")
    return out


def _load_clinvar_gene_level(wb) -> tuple[dict[str, str], dict[str, str]]:
    ws = wb["ClinVar_gene_level"]
    uid_map: dict[str, str] = {}
    mech_raw: dict[str, list[str]] = {}
    # A gene can span multiple rows; collect distinct non-empty UniProt IDs so a
    # conflict is surfaced rather than silently resolved by last-write-wins.
    uid_seen: dict[str, set[str]] = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if len(row) < 10:
            continue
        gene, uid, mech = row[0], row[1], row[9]
        if gene:
            if uid:
                uid_seen.setdefault(gene, set()).add(uid)
            uid_map.setdefault(gene, uid or "")
            if uid and not uid_map[gene]:
                uid_map[gene] = uid  # backfill if the first row had no UniProt ID
            mech_raw.setdefault(gene, []).append(mech or "Unknown")
    conflicts = {g: sorted(ids) for g, ids in uid_seen.items() if len(ids) > 1}
    if conflicts:
        examples = list(conflicts.items())[:5]
        print(
            f"WARNING: {len(conflicts)} genes map to multiple UniProt IDs in "
            f"ClinVar_gene_level; kept the first. Examples: {examples}"
        )
    mech_map: dict[str, str] = {
        gene: CV_MECH_MAP.get(Counter(mechs).most_common(1)[0][0], "Unknown")
        for gene, mechs in mech_raw.items()
    }
    print(f"ClinVar_gene_level: {len(uid_map)} genes")
    return uid_map, mech_map


def _load_g2p(path: Path) -> dict[str, str]:
    """Gene → mechanism from G2P (definitive/strong confidence only).

    Tiebreaking: if a gene has conflicting mechanisms, use only its 'definitive'
    entries. If that resolves to a single mechanism, accept it. If the conflict
    persists even at definitive confidence, exclude the gene.
    """
    df = pd.read_csv(path)
    filtered = df[
        df["confidence"].isin(G2P_CONFIDENCE_KEEP)
        & df["molecular mechanism"].isin(G2P_MECH_MAP)
    ].copy()
    filtered["mech_short"] = filtered["molecular mechanism"].map(G2P_MECH_MAP)

    out = {}
    skipped = []
    for gene, group in filtered.groupby("gene symbol"):
        unique = group["mech_short"].unique()
        if len(unique) == 1:
            out[gene] = unique[0]
            continue
        definitive = group[group["confidence"] == "definitive"]["mech_short"].unique()
        if len(definitive) == 1:
            out[gene] = definitive[0]
        else:
            skipped.append(gene)

    if skipped:
        print(
            f"G2P: excluded {len(skipped)} genes with unresolvable conflicting mechanisms: {sorted(skipped)}"
        )
    print(f"G2P (definitive/strong, unambiguous): {len(out)} genes")
    return out


def main() -> None:
    for path in [XLSX, G2P_CSV]:
        if not path.exists():
            raise FileNotFoundError(f"Required input not found: {path}")

    print(f"Loading {XLSX.name}")
    wb = openpyxl.load_workbook(XLSX, read_only=True)
    try:
        func_mechs = _load_functional_protein_class(wb)
        cv_uid, cv_mechs = _load_clinvar_gene_level(wb)
    finally:
        wb.close()
    g2p_best = _load_g2p(G2P_CSV)

    all_gera_genes = set(func_mechs) | set(cv_mechs)
    print(f"Total Gerasimavicius genes (union of sheets): {len(all_gera_genes)}")

    rows: list[dict] = []
    emitted_genes: set[str] = set()

    for gene in sorted(all_gera_genes):
        mech = func_mechs.get(gene)
        if not mech or mech == "Unknown":
            mech = cv_mechs.get(gene)
        uid = cv_uid.get(gene, "")
        g2p_mech = g2p_best.get(gene)

        if mech and mech != "Unknown":
            disagrees = g2p_mech if (g2p_mech and g2p_mech != mech) else None
            rows.append({
                "gene": gene,
                "mechanism": mech,
                "uniprot_id": uid,
                "source": "gerasimavicius",
                "g2p_disagrees": disagrees or "",
            })
            emitted_genes.add(gene)
        elif g2p_mech:
            rows.append({
                "gene": gene,
                "mechanism": g2p_mech,
                "uniprot_id": uid,
                "source": "g2p",
                "g2p_disagrees": "",
            })
            emitted_genes.add(gene)

    for gene in sorted(g2p_best):
        if gene not in emitted_genes:
            rows.append({
                "gene": gene,
                "mechanism": g2p_best[gene],
                "uniprot_id": "",
                "source": "g2p",
                "g2p_disagrees": "",
            })
            emitted_genes.add(gene)

    present_genes = {r["gene"] for r in rows}
    for alias in LEGACY_GENE_ALIASES:
        if alias["gene"] not in present_genes:
            rows.append(alias)

    rows.sort(key=lambda r: r["gene"])

    GENE_LIST_TSV.parent.mkdir(parents=True, exist_ok=True)
    fieldnames = ["gene", "mechanism", "uniprot_id", "source", "g2p_disagrees"]
    with open(GENE_LIST_TSV, "w", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames, delimiter="\t")
        writer.writeheader()
        writer.writerows(rows)

    src_counts = Counter(r["source"] for r in rows)
    mech_counts = Counter(r["mechanism"] for r in rows)
    n_disagree = sum(1 for r in rows if r["g2p_disagrees"])
    print(f"Wrote {len(rows)} genes to {GENE_LIST_TSV}")
    print(f"  source: {dict(src_counts)}")
    print(f"  mechanism: {dict(mech_counts)}")
    print(f"  g2p_disagrees: {n_disagree} genes")


if __name__ == "__main__":
    main()
