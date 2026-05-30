"""
Build the canonical gene universe for downstream feature scripts.

Two pipeline steps, run in order:

  Step 1 — gene-list  (run before variant fetching)
    Merges Gerasimavicius et al. and G2P datasets → merged_gene_list.tsv.
    Inputs : data/downloads/DiseaseMech_Stability_VEPS.xlsx
             data/downloads/AllG2P.csv
    Output : data/merged_gene_list.tsv

  Step 2 — universe  (run after fetch_annotations --step pfam)
    Filters merged_gene_list.tsv to genes with a Pfam family assignment.
    The output gene_universe.tsv is the canonical aligned row order for all
    feature matrices (proteome_features_aligned.npy, etc.).
    Inputs : data/merged_gene_list.tsv
             data/pfam_families.json
    Output : data/gene_universe.tsv
             Columns: gene, mechanism, uniprot_id, source, g2p_disagrees, pfam_family

Usage:
    python -m esm2_mechanism.fetch_data.build_gene_universe --step gene-list
    python -m esm2_mechanism.fetch_data.build_gene_universe --step universe
"""

from __future__ import annotations

import argparse
import csv
import functools
import json
from collections import Counter
from pathlib import Path

import openpyxl
import pandas as pd

from esm2_mechanism.utils_paths import DATA_DIR

print = functools.partial(print, flush=True)

# ---------------------------------------------------------------------------
# Paths
# ---------------------------------------------------------------------------
XLSX = DATA_DIR / "downloads" / "DiseaseMech_Stability_VEPS.xlsx"
G2P_CSV = DATA_DIR / "downloads" / "AllG2P.csv"
MERGED_GENE_LIST = DATA_DIR / "merged_gene_list.tsv"
PFAM_FAMILIES = DATA_DIR / "pfam_families.json"
GENE_UNIVERSE = DATA_DIR / "gene_universe.tsv"

# ---------------------------------------------------------------------------
# Step 1 constants
# ---------------------------------------------------------------------------
G2P_MECH_MAP = {
    "loss of function": "LOF",
    "gain of function": "GOF",
    "dominant negative": "DN",
}
G2P_CONFIDENCE_KEEP = {"definitive", "strong"}

LEGACY_GENE_ALIASES: list[dict] = [
    {
        "gene": "C12ORF65",
        "mechanism": "AR",
        "uniprot_id": "Q9H3J6",
        "source": "gerasimavicius",
        "g2p_disagrees": "",
    },
    {
        "gene": "C19ORF12",
        "mechanism": "LOF",
        "uniprot_id": "",
        "source": "g2p",
        "g2p_disagrees": "",
    },
]

CV_MECH_MAP = {
    "AR, Het": "AR",
    "AR": "AR",
    "HI": "HI",
    "GOF": "GOF",
    "DN": "DN",
    "Unknown": "Unknown",
}


# ---------------------------------------------------------------------------
# Step 1 helpers
# ---------------------------------------------------------------------------
def _load_functional_protein_class(wb) -> dict[str, str]:
    ws = wb["Functional_protein_class"]
    out: dict[str, str] = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        gene, _inh, mech = row[0], row[1], row[2]
        if gene:
            out[gene] = mech or "Unknown"
    print(f"Functional_protein_class: {len(out)} genes")
    return out


def _load_clinvar_gene_level(wb) -> tuple[dict[str, str], dict[str, str]]:
    ws = wb["ClinVar_gene_level"]
    uid_map: dict[str, str] = {}
    mech_raw: dict[str, list[str]] = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if len(row) < 10:
            continue
        gene, uid, mech = row[0], row[1], row[9]
        if gene:
            uid_map[gene] = uid or ""
            mech_raw.setdefault(gene, []).append(mech or "Unknown")
    mech_map: dict[str, str] = {
        gene: CV_MECH_MAP.get(Counter(mechs).most_common(1)[0][0], "Unknown")
        for gene, mechs in mech_raw.items()
    }
    print(f"ClinVar_gene_level: {len(uid_map)} genes")
    return uid_map, mech_map


def _load_g2p(path: Path) -> dict[str, str]:
    """Gene → mechanism from G2P (definitive/strong confidence only).

    Tiebreaking: if a gene has conflicting mechanisms, use only its 'definitive'
    entries. If that resolves to a single mechanism, accept it. If the conflict
    persists even at definitive confidence, exclude the gene.
    """
    df = pd.read_csv(path)
    filtered = df[
        df["confidence"].isin(G2P_CONFIDENCE_KEEP)
        & df["molecular mechanism"].isin(G2P_MECH_MAP)
    ].copy()
    filtered["mech_short"] = filtered["molecular mechanism"].map(G2P_MECH_MAP)

    out = {}
    skipped = []
    for gene, group in filtered.groupby("gene symbol"):
        unique = group["mech_short"].unique()
        if len(unique) == 1:
            out[gene] = unique[0]
            continue
        definitive = group[group["confidence"] == "definitive"]["mech_short"].unique()
        if len(definitive) == 1:
            out[gene] = definitive[0]
        else:
            skipped.append(gene)

    if skipped:
        print(
            f"G2P: excluded {len(skipped)} genes with unresolvable conflicting mechanisms: {sorted(skipped)}"
        )
    print(f"G2P (definitive/strong, unambiguous): {len(out)} genes")
    return out


def build(xlsx_path: Path, g2p_path: Path, out_path: Path) -> None:
    print(f"Loading {xlsx_path.name}")
    wb = openpyxl.load_workbook(xlsx_path, read_only=True)
    try:
        func_mechs = _load_functional_protein_class(wb)
        cv_uid, cv_mechs = _load_clinvar_gene_level(wb)
    finally:
        wb.close()
    g2p_best = _load_g2p(g2p_path)

    all_gera_genes = set(func_mechs) | set(cv_mechs)
    print(f"Total Gerasimavicius genes (union of sheets): {len(all_gera_genes)}")

    rows: list[dict] = []
    emitted_genes: set[str] = set()

    for gene in sorted(all_gera_genes):
        mech = func_mechs.get(gene)
        if not mech or mech == "Unknown":
            mech = cv_mechs.get(gene)
        uid = cv_uid.get(gene, "")
        g2p_mech = g2p_best.get(gene)

        if mech and mech != "Unknown":
            disagrees = g2p_mech if (g2p_mech and g2p_mech != mech) else None
            rows.append(
                {
                    "gene": gene,
                    "mechanism": mech,
                    "uniprot_id": uid,
                    "source": "gerasimavicius",
                    "g2p_disagrees": disagrees or "",
                }
            )
            emitted_genes.add(gene)
        elif g2p_mech:
            rows.append(
                {
                    "gene": gene,
                    "mechanism": g2p_mech,
                    "uniprot_id": uid,
                    "source": "g2p",
                    "g2p_disagrees": "",
                }
            )
            emitted_genes.add(gene)

    for gene in sorted(g2p_best):
        if gene not in emitted_genes:
            rows.append(
                {
                    "gene": gene,
                    "mechanism": g2p_best[gene],
                    "uniprot_id": "",
                    "source": "g2p",
                    "g2p_disagrees": "",
                }
            )
            emitted_genes.add(gene)

    present_genes = {r["gene"] for r in rows}
    for alias in LEGACY_GENE_ALIASES:
        if alias["gene"] not in present_genes:
            rows.append(alias)

    rows.sort(key=lambda r: r["gene"])

    out_path.parent.mkdir(parents=True, exist_ok=True)
    fieldnames = ["gene", "mechanism", "uniprot_id", "source", "g2p_disagrees"]
    with open(out_path, "w", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames, delimiter="\t")
        writer.writeheader()
        writer.writerows(rows)

    src_counts = Counter(r["source"] for r in rows)
    mech_counts = Counter(r["mechanism"] for r in rows)
    n_disagree = sum(1 for r in rows if r["g2p_disagrees"])
    print(f"Wrote {len(rows)} genes to {out_path}")
    print(f"  source: {dict(src_counts)}")
    print(f"  mechanism: {dict(mech_counts)}")
    print(f"  g2p_disagrees: {n_disagree} genes")


# ---------------------------------------------------------------------------
# Step 2 — gene_universe.tsv
# ---------------------------------------------------------------------------
def _build_gene_universe(merged_path: Path, pfam_path: Path, out_path: Path) -> None:
    """Filter merged_gene_list to Pfam-annotated genes → gene_universe.tsv."""
    with open(pfam_path) as f:
        pfam: dict[str, str | None] = json.load(f)
    n_annotated = sum(1 for v in pfam.values() if v is not None)
    print(f"Pfam families loaded: {len(pfam)} genes, {n_annotated} annotated")

    rows_in: list[dict] = []
    with open(merged_path) as f:
        reader = csv.DictReader(f, delimiter="\t")
        for row in reader:
            rows_in.append(row)
    print(f"merged_gene_list: {len(rows_in)} genes")

    rows_out: list[dict] = []
    dropped: list[str] = []
    for row in rows_in:
        gene = row["gene"]
        pfam_family = pfam.get(gene)
        if pfam_family is None:
            dropped.append(gene)
            continue
        rows_out.append({**row, "pfam_family": pfam_family})

    if dropped:
        print(
            f"Dropped {len(dropped)} genes with no Pfam annotation: {sorted(dropped)}"
        )
    print(f"gene_universe: {len(rows_out)} genes retained")

    fieldnames = [
        "gene",
        "mechanism",
        "uniprot_id",
        "source",
        "g2p_disagrees",
        "pfam_family",
    ]
    with open(out_path, "w", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames, delimiter="\t")
        writer.writeheader()
        writer.writerows(rows_out)
    print(f"Wrote {out_path}")


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------
def main_gene_list() -> None:
    for path in [XLSX, G2P_CSV]:
        if not path.exists():
            raise FileNotFoundError(f"Required input not found: {path}")
    build(XLSX, G2P_CSV, MERGED_GENE_LIST)


def main_universe() -> None:
    for path in [MERGED_GENE_LIST, PFAM_FAMILIES]:
        if not path.exists():
            raise FileNotFoundError(f"Required input not found: {path}")
    _build_gene_universe(MERGED_GENE_LIST, PFAM_FAMILIES, GENE_UNIVERSE)


def main() -> None:
    parser = argparse.ArgumentParser(
        description="Build gene universe files. See module docstring for step order."
    )
    parser.add_argument(
        "--step",
        choices=["gene-list", "universe"],
        required=True,
        help=(
            "gene-list: merge Gerasimavicius + G2P → merged_gene_list.tsv  |  "
            "universe: filter to Pfam-annotated genes → gene_universe.tsv"
        ),
    )
    args = parser.parse_args()

    if args.step == "gene-list":
        main_gene_list()
    else:
        main_universe()


if __name__ == "__main__":
    main()
