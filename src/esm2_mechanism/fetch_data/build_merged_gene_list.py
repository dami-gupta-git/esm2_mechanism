"""
Build merged_gene_list.tsv from Gerasimavicius and G2P datasets.

Sources
-------
Gerasimavicius  data/downloads/DiseaseMech_Stability_VEPS.xlsx
G2P             data/downloads/AllG2P.csv

Logic
-----
1. Gerasimavicius genes (non-Unknown mechanism) → source=gerasimavicius
   - Mechanism from Functional_protein_class sheet; fallback to ClinVar_gene_level
   - UniProt from ClinVar_gene_level sheet
   - g2p_disagrees filled when G2P (definitive/strong) disagrees

2. All other genes with definitive/strong G2P entries → source=g2p
   - Includes genes not in Gerasimavicius AND genes in Gerasimavicius with Unknown mechanism

3. Genes with only Unknown mechanism in Gerasimavicius AND no G2P entry are excluded.

G2P mechanism mapping
---------------------
  loss of function            → LOF
  gain of function            → GOF
  dominant negative           → DN
  undetermined / other        → excluded

Output columns: gene, mechanism, uniprot_id, source, g2p_disagrees

Usage
-----
    python -m esm2_mechanism.fetch_data.build_merged_gene_list
"""

import csv
import functools
from collections import Counter
from pathlib import Path

import pandas as pd
import openpyxl

from esm2_mechanism.utils_paths import DATA_DIR

print = functools.partial(print, flush=True)

XLSX = DATA_DIR / "downloads" / "DiseaseMech_Stability_VEPS.xlsx"
G2P_CSV = DATA_DIR / "downloads" / "AllG2P.csv"
OUT_PATH = DATA_DIR / "merged_gene_list.tsv"

G2P_MECH_MAP = {
    "loss of function": "LOF",
    "gain of function": "GOF",
    "dominant negative": "DN",
}
G2P_CONFIDENCE_KEEP = {"definitive", "strong"}

# Legacy gene names present in the original merged list that no longer appear under
# these symbols in either source dataset (renamed in HGNC after the manual merge).
# Kept verbatim so downstream scripts that depend on these identifiers still work.
LEGACY_GENE_ALIASES: list[dict] = [
    {"gene": "C12ORF65", "mechanism": "AR",  "uniprot_id": "Q9H3J6", "source": "gerasimavicius", "g2p_disagrees": ""},
    {"gene": "C19ORF12", "mechanism": "LOF", "uniprot_id": "",        "source": "g2p",            "g2p_disagrees": ""},
]

# ClinVar_gene_level Disease_mechanism → canonical label
CV_MECH_MAP = {
    "AR, Het": "AR",
    "AR": "AR",
    "HI": "HI",
    "GOF": "GOF",
    "DN": "DN",
    "Unknown": "Unknown",
}


def _load_functional_protein_class(wb) -> dict[str, str]:
    """Gene → mechanism from Functional_protein_class sheet (most curated)."""
    ws = wb["Functional_protein_class"]
    out: dict[str, str] = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        gene, _inh, mech = row[0], row[1], row[2]
        if gene:
            out[gene] = mech or "Unknown"
    print(f"Functional_protein_class: {len(out)} genes")
    return out


def _load_clinvar_gene_level(wb) -> tuple[dict[str, str], dict[str, str]]:
    """Gene → (uniprot_id, mechanism) from ClinVar_gene_level sheet."""
    ws = wb["ClinVar_gene_level"]
    uid_map: dict[str, str] = {}
    mech_raw: dict[str, list[str]] = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        gene, uid, mech = row[0], row[1], row[9]
        if gene:
            uid_map[gene] = uid or ""
            mech_raw.setdefault(gene, []).append(mech or "Unknown")

    mech_map: dict[str, str] = {
        gene: CV_MECH_MAP.get(Counter(mechs).most_common(1)[0][0], "Unknown")
        for gene, mechs in mech_raw.items()
    }
    print(f"ClinVar_gene_level: {len(uid_map)} genes")
    return uid_map, mech_map


def _load_g2p(path: Path) -> dict[str, str]:
    """Gene → mechanism from G2P (definitive/strong confidence only)."""
    df = pd.read_csv(path)
    filtered = df[
        df["confidence"].isin(G2P_CONFIDENCE_KEEP)
        & df["molecular mechanism"].isin(G2P_MECH_MAP)
    ].copy()
    filtered["mech_short"] = filtered["molecular mechanism"].map(G2P_MECH_MAP)
    out = (
        filtered.groupby("gene symbol")["mech_short"]
        .agg(lambda s: s.mode()[0])
        .to_dict()
    )
    print(f"G2P (definitive/strong): {len(out)} genes")
    return out


def build(xlsx_path: Path, g2p_path: Path, out_path: Path) -> None:
    print(f"Loading {xlsx_path.name}")
    wb = openpyxl.load_workbook(xlsx_path, read_only=True)
    try:
        func_mechs = _load_functional_protein_class(wb)
        cv_uid, cv_mechs = _load_clinvar_gene_level(wb)
    finally:
        wb.close()
    g2p_best = _load_g2p(g2p_path)

    all_gera_genes = set(func_mechs) | set(cv_mechs)
    print(f"Total Gerasimavicius genes (union of sheets): {len(all_gera_genes)}")

    rows: list[dict] = []
    emitted_genes: set[str] = set()

    for gene in sorted(all_gera_genes):
        # Best mechanism: prefer Functional_protein_class, fallback to ClinVar
        mech = func_mechs.get(gene)
        if not mech or mech == "Unknown":
            mech = cv_mechs.get(gene)

        uid = cv_uid.get(gene, "")
        g2p_mech = g2p_best.get(gene)

        if mech and mech != "Unknown":
            disagrees = g2p_mech if (g2p_mech and g2p_mech != mech) else None
            rows.append({
                "gene": gene,
                "mechanism": mech,
                "uniprot_id": uid,
                "source": "gerasimavicius",
                "g2p_disagrees": disagrees or "",
            })
            emitted_genes.add(gene)
        elif g2p_mech:
            # Unknown mechanism in Gerasimavicius — use G2P instead
            rows.append({
                "gene": gene,
                "mechanism": g2p_mech,
                "uniprot_id": uid,
                "source": "g2p",
                "g2p_disagrees": "",
            })
            emitted_genes.add(gene)
        # else: Unknown in both sources — excluded

    # Add G2P genes not covered by Gerasimavicius at all
    for gene in sorted(g2p_best):
        if gene not in emitted_genes:
            rows.append({
                "gene": gene,
                "mechanism": g2p_best[gene],
                "uniprot_id": "",
                "source": "g2p",
                "g2p_disagrees": "",
            })

    # Inject legacy aliases unless a current symbol already covers them
    present_genes = {r["gene"] for r in rows}
    for alias in LEGACY_GENE_ALIASES:
        if alias["gene"] not in present_genes:
            rows.append(alias)

    rows.sort(key=lambda r: r["gene"])

    out_path.parent.mkdir(parents=True, exist_ok=True)
    fieldnames = ["gene", "mechanism", "uniprot_id", "source", "g2p_disagrees"]
    with open(out_path, "w", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames, delimiter="\t")
        writer.writeheader()
        writer.writerows(rows)

    src_counts = Counter(r["source"] for r in rows)
    mech_counts = Counter(r["mechanism"] for r in rows)
    n_disagree = sum(1 for r in rows if r["g2p_disagrees"])
    print(f"Wrote {len(rows)} genes to {out_path}")
    print(f"  source: {dict(src_counts)}")
    print(f"  mechanism: {dict(mech_counts)}")
    print(f"  g2p_disagrees: {n_disagree} genes")


def main() -> None:
    for path in [XLSX, G2P_CSV]:
        if not path.exists():
            raise FileNotFoundError(f"Required input not found: {path}")
    build(XLSX, G2P_CSV, OUT_PATH)


if __name__ == "__main__":
    main()
